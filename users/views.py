# users/views.py
import datetime
import os
import PyPDF2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import hashlib
from os.path import basename
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from .forms import UploadFileForm, CustomPasswordResetForm
from .models import UploadedFile
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import PasswordResetView, PasswordResetDoneView, PasswordResetConfirmView, PasswordResetCompleteView
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.core.files.base import ContentFile
from django.core.exceptions import PermissionDenied
from django.core.mail import EmailMessage
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from urllib.parse import quote
from django.utils.text import get_valid_filename
from django.utils.encoding import force_bytes
from django.views.decorators.cache import never_cache
from django.http import Http404, FileResponse, HttpResponseRedirect, HttpResponse
from django import forms
from django.template.loader import render_to_string
from django.contrib.sites.shortcuts import get_current_site
import mimetypes
from mimetypes import guess_type


def home(request):
    if not request.user.is_authenticated:
        return redirect('login_view')
    return render(request, 'home.html')

class CustomAuthenticationForm(AuthenticationForm):
    username = forms.CharField(label="Utilisateur :")
    password = forms.CharField(label="Mot de passe :", widget=forms.PasswordInput)

def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')  # Redirect to home page
            else:
                messages.error(request,"Invalid username or password.")
        else:
            messages.error(request,"Invalid username or password.")
    form = CustomAuthenticationForm()
    return render(request = request, template_name = "registration/login.html", context={"form":form})

class CustomPasswordResetView(PasswordResetView):
    form_class = CustomPasswordResetForm
    template_name = 'registration/password_reset_form.html'
    success_url = 'password_reset_done'
    from_email = 'system@controle-course-au-large.fr'

    def form_valid(self, form):
        opts = {
            'use_https': self.request.is_secure(),
            'token_generator': default_token_generator,
            'from_email': getattr(self, 'from_email', None),
            'email_template_name': 'registration/password_reset_email.html',
            'subject_template_name': 'registration/password_reset_subject.txt',
            'request': self.request,
            'html_email_template_name': getattr(self, 'html_email_template_name', None),
            'extra_email_context': getattr(self, 'extra_email_context', None),
        }
        self.save(**opts)
        return super().form_valid(form)

    def save(self, **kwargs):
        """
        Generates a one-use only link for resetting password and sends to the user.
        """
        user = User.objects.get(email=self.request.POST['email'])
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        current_site = get_current_site(self.request)
        mail_subject = 'Reset your password'
        message = render_to_string('registration/password_reset_email.html', {
            'user': user,
            'domain': current_site.domain,
            'uid': uid,
            'token': token,
        })
        email = EmailMessage(mail_subject, message, to=[user.email])
        email.send()
    
class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'registration/password_reset_done.html'
    
class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'registration/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')

class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'registration/password_reset_complete.html'
    success_url = reverse_lazy('home')

def logout_view(request):
    logout(request)
    # Redirigez vers une page de réussite.
    return HttpResponseRedirect('/login/')

def synchroniser_fichiers_avec_db(user):
    # Obtenir tous les fichiers de l'utilisateur depuis la base de données
    fichiers_db = UploadedFile.objects.filter(user=user)

    for fichier_db in fichiers_db:
        chemin_fichier = os.path.join(settings.MEDIA_ROOT, fichier_db.file.name)
        
        # Vérifier si le fichier existe dans le répertoire
        if not os.path.exists(chemin_fichier):
            # Si le fichier n'existe pas, supprimer l'entrée de la base de données
            fichier_db.delete()


@login_required
def user_list(request):
    if request.user.is_superuser:
        users = User.objects.all()
    else:
        users = User.objects.filter(username=request.user.username)
    return render(request, 'user_list.html', {'users': users})


def rename_pdf(file_path, uploaded_file_id, user_id):
    if not os.path.exists(file_path):
        print(f"Le fichier {file_path} n'existe pas.")
        return

    with open(file_path, 'rb') as pdf_file:
        reader = PyPDF2.PdfReader(pdf_file)
        if reader.is_encrypted:
            reader.decrypt('')
        fields = reader.get_fields()
        number = fields.get('Numéro de Voile').value if fields.get('Numéro de Voile') else 'unknown_number'
        name_of_the_bapteme = fields.get('Nom de baptême').value if fields.get('Nom de baptême') else 'no_bapteme_name'
        name_of_race = fields.get('Nom du bateau').value if fields.get('Nom du bateau') else 'no_boat_name'
        name_of_the_skipper = fields.get('Soussigné').value if fields.get('Soussigné') else 'no_name'

    new_filename = f"{number}_{name_of_the_bapteme}_{name_of_race}_{name_of_the_skipper}.pdf".replace(" ", "_")
    new_filepath = os.path.join(settings.MEDIA_ROOT, f'user_{user_id}', new_filename)

    if not os.path.exists(os.path.dirname(new_filepath)):
        os.makedirs(os.path.dirname(new_filepath))

    os.rename(file_path, new_filepath)

    uploaded_file = get_object_or_404(UploadedFile, id=uploaded_file_id)
    uploaded_file.file = os.path.join(f'user_{user_id}', new_filename)  # Sauvegarder le chemin relatif
    uploaded_file.save()

    print(f"Fichier renommé sous le nom : {new_filename}")

def get_file_id(file_name, user_id):
    try:
        file = UploadedFile.objects.get(original_name=file_name, user_id=user_id)
        return file.id
    except UploadedFile.DoesNotExist:
        return None


@never_cache
@login_required
def upload_page(request, user_id):
    user = get_object_or_404(User, id=user_id)
    file_path = os.path.join(settings.MEDIA_ROOT, f'user_{user_id}')

    if not request.user.is_superuser and request.user != user:
        raise PermissionDenied

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = form.save(commit=False)
            uploaded_file.user = user
            uploaded_file.original_name = request.FILES['file'].name
            uploaded_file.save()
            return redirect('upload_page', user_id=user_id)
    else:
        form = UploadFileForm()
        
    extraire_donnees_pdf()

    if not os.path.exists(file_path):
        os.makedirs(file_path)

    files_in_directory = os.listdir(file_path)
    files = [{'name': file, 'path': os.path.join(f'user_{user_id}', quote(file))} for file in files_in_directory]

    return render(request, 'upload_page.html', {'files': files, 'user': user, 'form': form})

def extraire_donnees_pdf():
    # Préparer une liste pour stocker les données extraites
    donnees_extraites = []
    chemin_base = settings.MEDIA_ROOT
    # Parcourir tous les fichiers de tous les utilisateurs
    for dossier, sous_dossiers, fichiers in os.walk(chemin_base):
        if dossier == chemin_base:
            continue
        for nom_fichier in fichiers:
            chemin_fichier = os.path.join(dossier, nom_fichier)
            if chemin_fichier.endswith('.pdf'):
                try:
                    with open(chemin_fichier, 'rb') as pdf_file:
                        reader = PyPDF2.PdfReader(pdf_file)
                        if reader.is_encrypted:
                            reader.decrypt('')
                        fields = reader.get_fields()
                        # Extraire les données nécessaires
                        donnees = {
                            'Numéro de course': fields.get('1').value if fields.get('1') else 'Inconnu',
                            'Nom de course': fields.get('2').value if fields.get('2') else 'Inconnu',
                            'Nom de baptême': fields.get('3').value if fields.get('3') else 'Inconnu',
                            'Skipper': fields.get('4').value if fields.get('4') else 'Inconnu',
                            'Responsable du bateau pour les contrôles': fields.get('5').value if fields.get('5') else 'Inconnu',
                            'Téléphone': fields.get('6').value if fields.get('6') else 'Inconnu',
                            'MMSI': fields.get('7').value if fields.get('7') else 'Inconnu',
                            'N° de série': fields.get('8').value if fields.get('8') else 'Inconnu',
                            'prochaine révision': fields.get('9').value if fields.get('9') else 'Inconnu',
                            'Marque gilet skipper': fields.get('10').value if fields.get('10') else 'Inconnu',
                            'Couleur gilet skipper': fields.get('11').value if fields.get('11') else 'Inconnu',
                            'Nom inscrit sur le gilet skipper': fields.get('12').value if fields.get('12') else 'Inconnu',
                            'Marque gilet Spare': fields.get('13').value if fields.get('13') else 'Inconnu',
                            'Couleur gilet Spare': fields.get('14').value if fields.get('14') else 'Inconnu',
                            'Nom inscrit sur le gilet Spare': fields.get('15').value if fields.get('15') else 'Inconnu',
                            'Balise Personelle AIS': fields.get('16').value if fields.get('16') else 'Inconnu',
                            '2ème Balise Personelle AIS': fields.get('17').value if fields.get('17') else 'Inconnu',
                            'PLB': fields.get('18').value if fields.get('18') else 'Inconnu',
                            '2ème PLB': fields.get('19').value if fields.get('19') else 'Inconnu',
                            'Numéro combinaison de survie': fields.get('20').value if fields.get('20') else 'Inconnu',
                            'Couleur combinaison de survie': fields.get('21').value if fields.get('21') else 'Inconnu',
                            'Marque combinaison de survie': fields.get('22').value if fields.get('22') else 'Inconnu',
                            'Indicatif VHF': fields.get('23').value if fields.get('23') else 'Inconnu',
                            'Numéro Iridium': fields.get('24').value if fields.get('24') else 'Inconnu',
                            'N° EPIRB': fields.get('25').value if fields.get('25') else 'Inconnu',
                            'Date révision EPIRB': fields.get('26').value if fields.get('26') else 'Inconnu',
                            'Soussigné': fields.get('27').value if fields.get('27') else 'Inconnu',
                            'Date': fields.get('28').value if fields.get('28') else 'Inconnu',
                        }
                        
                        
                        def verifier_cases(donnees):
                            for key, value in donnees.items():
                                if "Case" in key:
                                    if value.value is None or value.value != "/Oui":
                                        return False
                            return True
                        result = verifier_cases(fields)
                        print(result)
                        donnees['Controle'] = 'OK' if result else 'Non OK'
                        donnees_extraites.append(donnees)
                            
                except Exception as e:
                    print(f"Erreur lors de la lecture du fichier {chemin_fichier}: {e}")
                

    # Créer un DataFrame avec les données extraites
    df = pd.DataFrame(donnees_extraites)

    # Sauvegarder le DataFrame dans un fichier Excel
    nom_fichier_excel = 'donnees_extraites.xlsx'
    chemin_complet = os.path.join(settings.MEDIA_ROOT, nom_fichier_excel)
    df.to_excel(chemin_complet, index=False)

    colorer_lignes_excel(chemin_complet)

    return chemin_complet

def colorer_lignes_excel(chemin_fichier_excel):
    wb = load_workbook(chemin_fichier_excel)
    ws = wb.active

    # Définir les couleurs
    fill_vert = PatternFill(start_color="CFFFB3", end_color="CFFFB3", fill_type="solid")
    fill_rouge = PatternFill(start_color="FF4B3E", end_color="FF4B3E", fill_type="solid")

    # Appliquer les couleurs aux lignes
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        couleur = row[-1].value  # La colonne 'Couleur' est la dernière colonne
        fill = fill_vert if couleur == 'OK' else fill_rouge
        for cell in row:
            cell.fill = fill

    wb.save(chemin_fichier_excel)




@login_required
def download(request, file_path):
    # Assurez-vous que l'utilisateur a le droit de télécharger ce fichier
    # Ici, vous pouvez ajouter une logique pour vérifier cela

    # Construire le chemin complet du fichier
    full_file_path = os.path.join(settings.MEDIA_ROOT, file_path)

    if not os.path.exists(full_file_path):
        raise Http404("Le fichier n'existe pas")

    with open(full_file_path, 'rb') as f:
        file_content = f.read()

    mime_type, _ = mimetypes.guess_type(full_file_path)
    response = HttpResponse(file_content, content_type=mime_type)
    response['Content-Disposition'] = f'attachment; filename="{quote(os.path.basename(full_file_path))}"'
    return response

@login_required
def delete_file(request, user_id, file_name):
    if not request.user.is_superuser and request.user.id != int(user_id):
        raise PermissionDenied

    file_path = os.path.join(settings.MEDIA_ROOT, f'user_{user_id}', quote(file_name))
    if os.path.exists(file_path):
        os.remove(file_path)
        UploadedFile.objects.filter(user_id=user_id, file=os.path.join(f'user_{user_id}', file_name)).delete()

    return redirect('upload_page', user_id=user_id)